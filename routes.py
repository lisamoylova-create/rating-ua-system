import os
import csv
import logging
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
from models_full import Company, SelectionBase, SelectionCompany, Ranking, RankingCompany, User
from app import db
from permissions import admin_required, manager_or_admin_required, upload_required, actualize_required, export_required, require_role
from sqlalchemy import desc, asc, text
from datetime import datetime
from data_processor_full import process_second_file
import pandas as pd

main = Blueprint('main', __name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

def is_company_actualized(actualized_value):
    """
    Універсальна функція для визначення чи актуалізована компанія
    Працює однаково в development і production
    """
    if not actualized_value:
        return False
    
    # Очищаємо значення 
    clean_value = str(actualized_value).strip().lower()
    
    # Список значень що означають "актуалізовано"
    actualized_values = ['так', 'yes', 'true', '1', 'актуалізовано', 'updated', 'ok']
    
    return clean_value in actualized_values

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def clean_text_value(value):
    """Clean text value for safe database storage"""
    if value is None or str(value).lower() in ['nan', 'none']:
        return None
    
    text = str(value).strip()
    if text.lower() == 'nan' or text == '':
        return None
    
    # Remove problematic characters
    text = text.replace('\x00', '').replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
    text = text.replace('\t', ' ')
    
    # Limit length to prevent database errors
    if len(text) > 500:
        text = text[:500]
    
    return text

def clean_numeric_value(value):
    """Clean numeric value for safe database storage"""
    if value is None or str(value).lower() in ['nan', 'none']:
        return None
    
    try:
        text = str(value).strip().replace(',', '').replace(' ', '').replace('-', '0')
        if text == '' or text.lower() == 'nan':
            return None
        return float(text)
    except:
        return None

def process_excel_data_optimized(file_path, file_type='basic'):
    """Optimized processing for large files up to 160K rows"""
    success_count = 0
    error_count = 0
    
    # Convert Excel to CSV first if needed
    csv_file_path = file_path.replace('.xlsx', '.csv').replace('.xls', '.csv')
    
    if file_path.endswith(('.xlsx', '.xls')):
        try:
            # Use openpyxl only for conversion to CSV
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                for row in ws.iter_rows(values_only=True):
                    cleaned_row = []
                    for cell in row:
                        if cell is not None:
                            cleaned_row.append(str(cell))
                        else:
                            cleaned_row.append('')
                    writer.writerow(cleaned_row)
            wb.close()
            
        except Exception as e:
            logging.error(f"Error converting Excel to CSV: {e}")
            return 0, 1
    else:
        csv_file_path = file_path
    
    # Process CSV file
    try:
        with open(csv_file_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            
            companies_to_create = []
            row_count = 0
            
            for row in reader:
                row_count += 1
                try:
                    # Get EDRPOU
                    edrpou = None
                    for key in ['ЄДРПОУ', 'Код ЄДРПОУ', 'edrpou']:
                        if key in row and row[key]:
                            edrpou = str(row[key]).strip()
                            break
                    
                    if not edrpou or edrpou == 'nan':
                        error_count += 1
                        continue
                    
                    # Collect company data
                    company_data = {
                        'edrpou': edrpou,
                        'name': (row.get('Назва', '') or row.get('Название компании', '') or row.get('Назва компанії', ''))[:500],
                        'kved_code': (row.get('КВЕД', '') or row.get('Код КВЕД', ''))[:20],
                        'region': (row.get('Регіон', '') or row.get('Область', ''))[:100]
                    }
                    
                    # Parse numeric fields safely
                    try:
                        personnel_str = row.get('Кількість працівників', '') or row.get('Персонал (2019 р.)', '') or row.get('Персонал (осіб)', '')
                        if personnel_str and personnel_str != 'nan':
                            company_data['personnel'] = int(float(personnel_str))
                    except:
                        company_data['personnel'] = None
                    
                    try:
                        revenue_str = row.get('Дохід', '') or row.get('Чистий дохід (тис. грн)', '')
                        if revenue_str and revenue_str != 'nan':
                            company_data['revenue'] = float(revenue_str)
                    except:
                        company_data['revenue'] = None
                    
                    companies_to_create.append(company_data)
                    
                    # Process in batches of 200 for large file efficiency
                    if len(companies_to_create) >= 200:
                        batch_success = process_simple_batch(companies_to_create)
                        success_count += batch_success
                        error_count += (len(companies_to_create) - batch_success)
                        companies_to_create = []
                        
                        if row_count % 2000 == 0:
                            logging.info(f"Processed {row_count} rows: {success_count} success, {error_count} errors")
                
                except Exception as e:
                    logging.error(f"Error processing row {row_count}: {e}")
                    error_count += 1
                    continue
            
            # Process remaining companies
            if companies_to_create:
                batch_success = process_simple_batch(companies_to_create)
                success_count += batch_success
                error_count += (len(companies_to_create) - batch_success)
        
        # Clean up CSV file if it was converted
        if csv_file_path != file_path:
            os.remove(csv_file_path)
            
    except Exception as e:
        logging.error(f"Error processing CSV: {e}")
        error_count += 100
    
    return success_count, error_count

def process_simple_batch(companies_data):
    """Efficient batch processing with bulk operations"""
    success_count = 0
    
    try:
        # Get all existing EDRPOU codes in this batch for efficient lookup
        edrpou_list = [data['edrpou'] for data in companies_data]
        existing_companies = {
            company.edrpou: company for company in 
            db.session.query(Company).filter(Company.edrpou.in_(edrpou_list)).all()
        }
        
        # Process companies
        for data in companies_data:
            try:
                edrpou = data['edrpou']
                
                if edrpou in existing_companies:
                    # Update existing company
                    company = existing_companies[edrpou]
                    if data['name']:
                        company.name = data['name']
                    if data['kved_code']:
                        company.kved_code = data['kved_code']
                    if data['region']:
                        company.region_name = data['region']
                    if data.get('personnel'):
                        company.personnel_2019 = data['personnel']
                    if data.get('revenue'):
                        company.revenue_2019 = data['revenue']
                else:
                    # Create new company
                    company = Company(
                        edrpou=edrpou,
                        name=data['name'][:500] if data['name'] else None,
                        kved_code=data['kved_code'][:20] if data['kved_code'] else None,
                        region_name=data['region'][:100] if data['region'] else None,
                        personnel_2019=data.get('personnel'),
                        revenue_2019=data.get('revenue'),
                        source='основний',
                        actualized='ні'
                    )
                    db.session.add(company)
                
                success_count += 1
                
            except Exception as e:
                logging.error(f"Error processing company {data['edrpou']}: {e}")
                continue
        
        # Single commit for entire batch
        db.session.commit()
        
    except Exception as e:
        logging.error(f"Error in batch processing: {e}")
        db.session.rollback()
        success_count = 0
    
    return success_count

def process_batch_to_database(batch_data):
    """Process a batch of companies using PostgreSQL UPSERT for maximum performance"""
    success_count = 0
    error_count = 0
    
    # Prepare values list for bulk upsert
    values_list = []
    
    for data in batch_data:
        try:
            edrpou = data['edrpou']
            name = data['name'] if data['name'] else None
            kved_code = data['kved_code'] if data['kved_code'] else None
            region = data['region'] if data['region'] else None
            
            personnel_val = None
            revenue_val = None
            
            if data['personnel']:
                try:
                    personnel_val = int(float(data['personnel']))
                except:
                    pass
            
            if data['revenue']:
                try:
                    revenue_val = float(data['revenue'])
                except:
                    pass
            
            values_list.append((edrpou, name, kved_code, personnel_val, revenue_val, region))
            
        except Exception as e:
            logging.error(f"Error preparing data for {data.get('edrpou', 'unknown')}: {e}")
            error_count += 1
            continue
    
    if not values_list:
        return 0, error_count
    
    try:
        # Use PostgreSQL UPSERT for bulk insert/update
        upsert_sql = """
            INSERT INTO companies (edrpou, name, kved_code, personnel_2019, revenue_2019, region_name, source, actualized, created_at, updated_at)
            VALUES %s
            ON CONFLICT (edrpou) 
            DO UPDATE SET 
                name = EXCLUDED.name,
                kved_code = EXCLUDED.kved_code,
                personnel_2019 = EXCLUDED.personnel_2019,
                revenue_2019 = EXCLUDED.revenue_2019,
                region_name = EXCLUDED.region_name,
                updated_at = CURRENT_TIMESTAMP
        """
        
        # Create VALUES clause
        values_str = ','.join([
            f"('{edrpou}', " +
            (f"'{name}'" if name else 'NULL') + ", " +
            (f"'{kved_code}'" if kved_code else 'NULL') + ", " +
            f"{personnel if personnel else 'NULL'}, {revenue if revenue else 'NULL'}, " +
            (f"'{region}'" if region else 'NULL') + ", 'основний', 'ні', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
            for edrpou, name, kved_code, personnel, revenue, region in values_list
        ])
        
        final_sql = upsert_sql.replace('%s', values_str)
        
        # Execute the bulk upsert
        db.session.execute(text(final_sql))
        db.session.commit()
        
        success_count = len(values_list)
        
    except Exception as e:
        logging.error(f"Error in bulk upsert: {e}")
        db.session.rollback()
        
        # Fallback: process one by one
        for edrpou, name, kved_code, personnel, revenue, region in values_list:
            try:
                simple_upsert = text("""
                    INSERT INTO companies (edrpou, name, kved_code, personnel_2019, revenue_2019, region_name, source, actualized, created_at, updated_at)
                    VALUES (:edrpou, :name, :kved_code, :personnel, :revenue, :region, 'основний', 'ні', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    ON CONFLICT (edrpou) 
                    DO UPDATE SET 
                        name = EXCLUDED.name,
                        kved_code = EXCLUDED.kved_code,
                        personnel_2019 = EXCLUDED.personnel_2019,
                        revenue_2019 = EXCLUDED.revenue_2019,
                        region_name = EXCLUDED.region_name,
                        updated_at = CURRENT_TIMESTAMP
                """)
                
                db.session.execute(simple_upsert, {
                    'edrpou': edrpou,
                    'name': name,
                    'kved_code': kved_code,
                    'personnel': personnel,
                    'revenue': revenue,
                    'region': region
                })
                db.session.commit()
                success_count += 1
                
            except Exception as e2:
                logging.error(f"Error in fallback upsert for {edrpou}: {e2}")
                db.session.rollback()
                error_count += 1
                continue
    
    return success_count, error_count

@main.route('/debug-stats')
@login_required 
def debug_stats():
    """Simple stats check"""
    try:
        total = db.session.execute(db.text("SELECT COUNT(*) FROM companies")).scalar()
        kved = db.session.execute(db.text("SELECT COUNT(DISTINCT kved_code) FROM companies WHERE kved_code IS NOT NULL AND kved_code != ''")).scalar()
        regions = db.session.execute(db.text("SELECT COUNT(DISTINCT region_name) FROM companies WHERE region_name IS NOT NULL AND region_name != ''")).scalar()
        
        sample = db.session.execute(db.text("SELECT edrpou, name, kved_code, region_name FROM companies LIMIT 5")).fetchall()
        
        return f"""<h1>Production Database Stats</h1>
        <p><strong>Total Companies:</strong> {total}</p>
        <p><strong>Unique KVED Codes:</strong> {kved}</p> 
        <p><strong>Unique Regions:</strong> {regions}</p>
        <h2>Sample Data:</h2>
        <table border='1'>
        <tr><th>EDRPOU</th><th>Name</th><th>KVED</th><th>Region</th></tr>
        {"".join(f"<tr><td>{row[0]}</td><td>{row[1][:50]}...</td><td>{row[2]}</td><td>{row[3]}</td></tr>" for row in sample)}
        </table>"""
    except Exception as e:
        return f"<h1>Error:</h1><p>{str(e)}</p>"

@main.route('/')
def index():
    if current_user.is_authenticated:
        try:
            # Use only Raw SQL queries - they work perfectly on production!
            total_companies = db.session.execute(db.text("SELECT COUNT(*) FROM companies")).scalar() or 0
            
            region_count_query = db.session.execute(
                db.text("SELECT COUNT(DISTINCT region_name) FROM companies WHERE region_name IS NOT NULL AND region_name != ''")
            ).scalar() or 0
            
            kved_count_query = db.session.execute(
                db.text("SELECT COUNT(DISTINCT kved_code) FROM companies WHERE kved_code IS NOT NULL AND kved_code != ''")
            ).scalar() or 0
            
            logging.info(f"Stats loaded: companies={total_companies}, regions={region_count_query}, kved={kved_count_query}")
            
            # Recent companies using working Raw SQL pattern
            recent_companies = []
            try:
                recent_rows = db.session.execute(
                    db.text("SELECT edrpou, name, region_name, kved_code, created_at FROM companies ORDER BY created_at DESC NULLS LAST LIMIT 5")
                ).fetchall()
                
                for row in recent_rows:
                    recent_companies.append(type('Company', (), {
                        'edrpou': row[0], 
                        'name': row[1], 
                        'region_name': row[2], 
                        'kved_code': row[3],
                        'created_at': row[4] if row[4] else type('datetime', (), {'strftime': lambda x: 'Невідомо'})()
                    })())
                    
                logging.info(f"Recent companies loaded: {len(recent_companies)}")
            except Exception as e:
                logging.error(f"Error getting recent companies: {e}")
            
            # KVED statistics - правильна логіка для відбору та актуалізації
            kved_stats = []
            try:
                # Перевіряємо чи є активні відбори
                active_selection = db.session.execute(
                    db.text("SELECT id FROM selection_bases WHERE is_active = true LIMIT 1")
                ).scalar()
                
                if active_selection:
                    # Якщо є активний відбір - рахуємо компанії в відборі + актуалізовані
                    kved_query = db.session.execute(
                        db.text("""
                            SELECT 
                                c.kved_code,
                                c.kved_description,
                                COUNT(*) as total_count,
                                COUNT(CASE WHEN sc.company_id IS NOT NULL THEN 1 END) as after_selection_count,
                                COUNT(CASE WHEN LOWER(TRIM(c.actualized)) IN ('так', 'yes', 'true', '1', 'актуалізовано', 'updated', 'ok') THEN 1 END) as really_actualized_count
                            FROM companies c
                            LEFT JOIN selection_companies sc ON c.id = sc.company_id 
                            WHERE c.kved_code IS NOT NULL AND c.kved_code != ''
                            GROUP BY c.kved_code, c.kved_description 
                            ORDER BY total_count DESC 
                            LIMIT 10
                        """)
                    ).fetchall()
                else:
                    # Якщо немає активного відбору - рахуємо тільки загальну кількість + актуалізовані
                    kved_query = db.session.execute(
                        db.text("""
                            SELECT 
                                kved_code,
                                kved_description,
                                COUNT(*) as total_count,
                                0 as after_selection_count,
                                COUNT(CASE WHEN LOWER(TRIM(actualized)) IN ('так', 'yes', 'true', '1', 'актуалізовано', 'updated', 'ok') THEN 1 END) as really_actualized_count
                            FROM companies 
                            WHERE kved_code IS NOT NULL AND kved_code != ''
                            GROUP BY kved_code, kved_description 
                            ORDER BY total_count DESC 
                            LIMIT 10
                        """)
                    ).fetchall()
                
                for row in kved_query:
                    after_selection_text = "Відбір не створений" if row[3] == 0 else str(row[3])
                    actualized_text = f"{row[4]}" if row[4] > 0 else "0"
                    
                    kved_stats.append({
                        'code': row[0],
                        'description': row[1][:50] + '...' if row[1] and len(str(row[1])) > 50 else (row[1] or 'Не вказано'),
                        'total_count': row[2],
                        'after_selection': row[3],  # Кількість після відбору (0 якщо відбору немає)
                        'selection_criteria': actualized_text  # Кількість реально актуалізованих
                    })
                    
                logging.info(f"KVED stats loaded: {len(kved_stats)} entries")
            except Exception as e:
                logging.error(f"Error getting KVED statistics: {e}")

            stats = {
                'total_companies': total_companies,
                'total_regions': region_count_query,
                'total_kved': kved_count_query,
                'recent_companies': recent_companies,
                'kved_statistics': kved_stats
            }
            
            # Debug logging
            logging.info(f"Dashboard stats: companies={total_companies}, regions={region_count_query}, kved={kved_count_query}, recent={len(recent_companies)}, kved_stats={len(kved_stats)}")
            
            return render_template('index.html', stats=stats)
        except Exception as e:
            logging.error(f"Error in index route: {e}")
            # Return basic template without stats
            return render_template('index.html', stats={
                'total_companies': 0,
                'total_regions': 0,
                'total_kved': 0,
                'recent_companies': [],
                'kved_statistics': []
            })
    else:
        return redirect(url_for('auth.login'))

@main.route('/upload', methods=['GET', 'POST'])
@login_required
@admin_required
def upload():
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if 'file' not in request.files:
            flash('Файл не обраний', 'danger')
            return redirect(request.url)
            
        file = request.files['file']
        if file.filename == '':
            flash('Файл не обраний', 'danger')
            return redirect(request.url)
            
        if file and file.filename and allowed_file(file.filename):
            try:
                filename = secure_filename(file.filename)
                
                # Save file to uploads directory
                uploads_dir = 'uploads'
                os.makedirs(uploads_dir, exist_ok=True)
                file_path = os.path.join(uploads_dir, filename)
                file.save(file_path)
                
                # Process file directly without pandas
                logging.info(f"Processing file: {filename}")
                
                if action == 'upload':
                    # Конвертація в CSV для подальшої обробки
                    if filename.endswith(('.xlsx', '.xls')):
                        csv_filename = filename.replace('.xlsx', '.csv').replace('.xls', '.csv')
                        csv_path = os.path.join('uploads', f"converted_{csv_filename}")
                        
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        ws = wb.active
                        
                        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                            writer = csv.writer(csvfile)
                            row_count = 0
                            for row in ws.iter_rows(values_only=True):
                                cleaned_row = []
                                for cell in row:
                                    if cell is not None:
                                        cleaned_row.append(str(cell))
                                    else:
                                        cleaned_row.append('')
                                writer.writerow(cleaned_row)
                                row_count += 1
                        
                        wb.close()
                        flash(f'Excel файл {filename} конвертовано в CSV. Знайдено {row_count} рядків. Файл збережено як {csv_filename} для подальшої обробки.', 'success')
                        
                    elif filename.endswith('.csv'):
                        # Скопіювати CSV файл
                        csv_path = os.path.join('uploads', f"imported_{filename}")
                        import shutil
                        shutil.copy2(file_path, csv_path)
                        
                        with open(file_path, 'r', encoding='utf-8') as f:
                            row_count = sum(1 for _ in f)
                        flash(f'CSV файл {filename} імпортовано. Знайдено {row_count} рядків. Готово для подальшої обробки.', 'success')
                        
                elif action == 'actualize':
                    # Конвертація файлу актуалізації в CSV для подальшої обробки (як у кроці 1)
                    if filename.endswith(('.xlsx', '.xls')):
                        csv_filename = filename.replace('.xlsx', '.csv').replace('.xls', '.csv')
                        csv_path = os.path.join('uploads', f"actualization_{csv_filename}")
                        
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        ws = wb.active
                        
                        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                            writer = csv.writer(csvfile)
                            row_count = 0
                            for row in ws.iter_rows(values_only=True):
                                cleaned_row = []
                                for cell in row:
                                    if cell is not None:
                                        cleaned_row.append(str(cell))
                                    else:
                                        cleaned_row.append('')
                                writer.writerow(cleaned_row)
                                row_count += 1
                        
                        wb.close()
                        flash(f'📁 Excel файл {filename} конвертовано для актуалізації. Знайдено {row_count} рядків. Файл збережено як actualization_{csv_filename} для подальшої обробки.', 'success')
                        
                    elif filename.endswith('.csv'):
                        # Скопіювати CSV файл для актуалізації
                        csv_path = os.path.join('uploads', f"actualization_{filename}")
                        import shutil
                        shutil.copy2(file_path, csv_path)
                        
                        with open(file_path, 'r', encoding='utf-8') as f:
                            row_count = sum(1 for _ in f)
                        
                        flash(f'📁 CSV файл {filename} підготовлено для актуалізації. Знайдено {row_count} рядків. Файл збережено як actualization_{filename}.', 'success')
                    else:
                        flash(f'❌ Недопустимий формат файлу для актуалізації. Використовуйте .xlsx, .xls або .csv файли.', 'danger')
                
                # Clean up the original uploaded file (converted files kept in uploads/)
                os.remove(file_path)
                return redirect(url_for('main.upload'))
                
            except Exception as e:
                logging.error(f"Error processing file: {str(e)}")
                flash(f'Помилка обробки файлу: {str(e)}', 'danger')
                return redirect(request.url)
        else:
            flash('Недозволений тип файлу. Дозволені: xlsx, xls, csv', 'danger')
            return redirect(request.url)
    
    return render_template('upload.html')

@main.route('/file-manager', methods=['GET', 'POST'])
@login_required
@admin_required
def file_manager():
    
    uploads_dir = 'uploads'
    os.makedirs(uploads_dir, exist_ok=True)
    
    if request.method == 'POST':
        action = request.form.get('action')
        filename = request.form.get('filename')
        
        if action == 'delete' and filename:
            file_path = os.path.join(uploads_dir, filename)
            if os.path.exists(file_path):
                os.remove(file_path)
                flash(f'Файл {filename} видалено.', 'success')
            else:
                flash(f'Файл {filename} не знайдено.', 'error')
                
        elif action == 'process' and filename:
            # Redirect to processing page
            return redirect(url_for('main.process_file', filename=filename))
            
        elif action == 'process_external' and filename:
            # Process with external script
            file_path = os.path.join(uploads_dir, filename)
            if os.path.exists(file_path) and filename.endswith('.csv'):
                try:
                    import subprocess
                    result = subprocess.run([
                        'python', 'process_large_csv.py', file_path
                    ], capture_output=True, text=True, timeout=300)
                    
                    if result.returncode == 0:
                        flash(f'Файл {filename} успішно оброблено зовнішнім скриптом. Перевірте консоль для деталей.', 'success')
                    else:
                        flash(f'Помилка обробки файлу {filename}: {result.stderr}', 'error')
                        
                except subprocess.TimeoutExpired:
                    flash(f'Обробка файлу {filename} перервана через таймаут (5 хвилин)', 'warning')
                except Exception as e:
                    flash(f'Помилка запуску зовнішнього скрипта: {str(e)}', 'error')
            else:
                flash(f'Файл {filename} не знайдено або це не CSV файл.', 'error')
                
        elif action == 'actualize_external' and filename:
            # Перенаправити на сторінку прогресу актуалізації
            file_path = os.path.join(uploads_dir, filename)
            if os.path.exists(file_path) and filename.startswith('actualization_') and filename.endswith('.csv'):
                return redirect(url_for('main.actualize_progress', filename=filename))
            else:
                flash(f'Файл {filename} не знайдено або це не файл актуалізації.', 'error')
                
        elif action == 'import_to_db' and filename:
            # Redirect to database import page
            return redirect(url_for('main.database_import', filename=filename))
            
        elif action == 'actualize_to_db' and filename:
            # Завантажити результат актуалізації в базу
            file_path = os.path.join(uploads_dir, filename)
            if os.path.exists(file_path) and filename.endswith('_actualized.csv'):
                # Оскільки актуалізація вже виконана зовнішнім скриптом, тільки показати результат
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        updated_count = sum(1 for _ in f) - 1  # -1 для заголовка
                    flash(f'✅ Актуалізація завершена! Оновлено {updated_count} компаній у базі даних.', 'success')
                except Exception as e:
                    flash(f'Помилка читання результату актуалізації: {str(e)}', 'error')
            else:
                flash(f'Файл результату актуалізації {filename} не знайдено.', 'error')
            
        elif action == 'bulk_import' and filename:
            # Швидке масове завантаження через COPY FROM
            file_path = os.path.join(uploads_dir, filename)
            if os.path.exists(file_path) and filename.endswith('_processed.csv'):
                try:
                    import subprocess
                    result = subprocess.run([
                        'python', 'bulk_import_optimized.py', file_path
                    ], capture_output=True, text=True, timeout=1800)  # 30 хвилин таймаут
                    
                    if result.returncode == 0:
                        flash(f'Масове завантаження {filename} завершено успішно за рекордний час!', 'success')
                    else:
                        flash(f'Помилка масового завантаження: {result.stderr}', 'error')
                        
                except subprocess.TimeoutExpired:
                    flash(f'Масове завантаження перервано через таймаут', 'warning')
                except Exception as e:
                    flash(f'Помилка запуску масового завантаження: {str(e)}', 'error')
            else:
                flash(f'Файл {filename} не знайдено або це не оброблений CSV файл.', 'error')
        
        return redirect(url_for('main.file_manager'))
    
    # Get list of files
    files = []
    try:
        for filename in os.listdir(uploads_dir):
            file_path = os.path.join(uploads_dir, filename)
            if os.path.isfile(file_path):
                stat = os.stat(file_path)
                files.append({
                    'name': filename,
                    'size_mb': round(stat.st_size / (1024 * 1024), 2),
                    'created_date': datetime.fromtimestamp(stat.st_ctime).strftime('%d.%m.%Y %H:%M')
                })
    except Exception as e:
        logging.error(f"Error listing files: {e}")
    
    # Sort by creation date (newest first)
    files.sort(key=lambda x: x['created_date'], reverse=True)
    
    return render_template('file_manager.html', files=files)

@main.route('/process-file/<filename>')
@login_required
@admin_required
def process_file(filename):
    """Show processing page for file"""
    
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, filename)
    
    if not os.path.exists(file_path):
        flash(f'Файл {filename} не знайдено.', 'error')
        return redirect(url_for('main.file_manager'))
    
    return render_template('process_progress.html', filename=filename)

@main.route('/actualize-file/<filename>')
@login_required 
@admin_required
def actualize_file(filename):
    """Show actualization processing page for file"""
    
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, filename)
    
    if not os.path.exists(file_path):
        flash(f'Файл {filename} не знайдено.', 'error')
        return redirect(url_for('main.file_manager'))
    
    return render_template('actualize_progress.html', filename=filename)

@main.route('/actualize-progress/<filename>')
@login_required 
@admin_required
def actualize_progress(filename):
    """Show actualization processing progress page"""
    
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, filename)
    
    if not os.path.exists(file_path):
        flash(f'Файл {filename} не знайдено.', 'error')
        return redirect(url_for('main.file_manager'))
    
    return render_template('actualize_processing.html', filename=filename)

@main.route('/api/actualize-csv/<filename>')
@login_required
@admin_required
def actualize_csv(filename):
    """Process actualization CSV file and return progress/results"""
    
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, filename)
    
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    if not filename.startswith('actualization_') or not filename.endswith('.csv'):
        return jsonify({'error': 'Invalid actualization file format'}), 400
    
    try:
        # Read actualization CSV file  
        df = pd.read_csv(file_path, encoding='utf-8')
        
        if df is None or df.empty:
            return jsonify({'error': 'Не вдалося прочитати файл актуалізації'}), 500
        
        total_lines = len(df)
        
        # Find EDRPOU column
        edrpou_col = None
        for col_name in df.columns:
            if any(keyword in col_name.upper() for keyword in ['ЄДРПОУ', 'EDRPOU']):
                edrpou_col = col_name
                break
        
        if not edrpou_col:
            return jsonify({'error': 'Колонка ЄДRПОУ не знайдена в файлі актуалізації'}), 400
        
        # Count valid EDRPOU
        valid_edrpou_count = 0
        sample_data = []
        
        for i, (index, row) in enumerate(df.iterrows()):
            edrpou = str(row[edrpou_col]).strip()
            if edrpou and edrpou != 'nan' and edrpou.isdigit():
                valid_edrpou_count += 1
                
                # Collect sample data (first 5 valid records)
                if len(sample_data) < 5:
                    sample_data.append({
                        'edrpou': edrpou,
                        'first_name': str(row.get('Ім\'я', '') or row.get('Имя', ''))[:50],
                        'last_name': str(row.get('Прізвище', '') or row.get('Фамилия', ''))[:50],
                        'work_phone': str(row.get('Робочий телефон', '') or row.get('Рабочий телефон', ''))[:30]
                    })
        
        return jsonify({
            'status': 'ready_for_processing',
            'total_lines': total_lines,
            'valid_edrpou_count': valid_edrpou_count, 
            'invalid_count': total_lines - valid_edrpou_count,
            'sample_data': sample_data,
            'message': f'Готово для актуалізації: {valid_edrpou_count} компаній з валідними ЄДRПОУ'
        })
        
    except Exception as e:
        logging.error(f"Error processing actualization file: {str(e)}")
        return jsonify({'error': f'Помилка обробки файлу актуалізації: {str(e)}'}), 500

@main.route('/api/actualize-start/<filename>')
@login_required
@admin_required
def actualize_start(filename):
    """Start actualization processing in background"""
    
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, filename)
    
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    if not filename.startswith('actualization_') or not filename.endswith('.csv'):
        return jsonify({'error': 'Invalid actualization file format'}), 400
    
    try:
        import subprocess
        import threading
        
        # Створити файл статусу
        status_file = file_path.replace('.csv', '_status.txt')
        with open(status_file, 'w', encoding='utf-8') as f:
            f.write('processing')
        
        def run_actualization():
            try:
                # Запустити зовнішній скрипт
                result = subprocess.run([
                    'python', 'process_actualization.py', file_path
                ], capture_output=True, text=True, timeout=900)
                
                if result.returncode == 0:
                    # Перевірити результат
                    result_filename = file_path.replace('.csv', '_actualized.csv')
                    if os.path.exists(result_filename):
                        with open(result_filename, 'r', encoding='utf-8') as f:
                            updated_count = sum(1 for _ in f) - 1
                        
                        # Записати успішний статус
                        with open(status_file, 'w', encoding='utf-8') as f:
                            f.write(f'completed:{updated_count}:{os.path.basename(result_filename)}')
                    else:
                        with open(status_file, 'w', encoding='utf-8') as f:
                            f.write('error:Результуючий файл не створено')
                else:
                    with open(status_file, 'w', encoding='utf-8') as f:
                        f.write(f'error:{result.stderr}')
                        
            except subprocess.TimeoutExpired:
                with open(status_file, 'w', encoding='utf-8') as f:
                    f.write('error:Таймаут 15 хвилин')
            except Exception as e:
                with open(status_file, 'w', encoding='utf-8') as f:
                    f.write(f'error:{str(e)}')
        
        # Запустити в окремому потоці
        thread = threading.Thread(target=run_actualization)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'status': 'started',
            'message': 'Обробка актуалізації запущена у фоновому режимі'
        })
        
    except Exception as e:
        logging.error(f"Error starting actualization: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': f'Помилка запуску: {str(e)}'
        })

@main.route('/api/actualize-status/<filename>')
@login_required
@admin_required
def actualize_status(filename):
    """Check actualization processing status"""
    
    uploads_dir = 'uploads'
    status_file = os.path.join(uploads_dir, filename.replace('.csv', '_status.txt'))
    
    if not os.path.exists(status_file):
        return jsonify({
            'status': 'not_started',
            'message': 'Обробка не запущена'
        })
    
    try:
        with open(status_file, 'r', encoding='utf-8') as f:
            status_content = f.read().strip()
        
        if status_content == 'processing':
            return jsonify({
                'status': 'processing',
                'message': 'Обробка триває...'
            })
        elif status_content.startswith('completed:'):
            parts = status_content.split(':')
            updated_count = int(parts[1]) if len(parts) > 1 else 0
            result_file = parts[2] if len(parts) > 2 else 'unknown'
            
            return jsonify({
                'status': 'completed',
                'message': f'Актуалізація завершена! Оновлено {updated_count} компаній.',
                'updated_count': updated_count,
                'result_file': result_file
            })
        elif status_content.startswith('error:'):
            error_msg = status_content[6:]  # Видалити 'error:'
            return jsonify({
                'status': 'error',
                'error': error_msg
            })
        else:
            return jsonify({
                'status': 'unknown',
                'message': 'Невідомий статус обробки'
            })
            
    except Exception as e:
        logging.error(f"Error checking status: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': f'Помилка перевірки статусу: {str(e)}'
        })

@main.route('/api/process-csv-simple/<filename>')
@login_required
def process_csv_simple(filename):
    """Simple CSV/Excel processing without database complications"""
    if not current_user.has_permission('edit'):
        return jsonify({'error': 'No permission'}), 403
    
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, filename)
    
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    try:
        # Read file based on type (Excel or CSV)
        df = None
        
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Handle Excel files with encoding issues
            for engine in ['openpyxl', 'xlrd', None]:
                try:
                    logging.info(f"API: Trying to read Excel with engine: {engine}")
                    if engine:
                        df = pd.read_excel(file_path, engine=engine)
                    else:
                        df = pd.read_excel(file_path)
                    
                    if df is not None and not df.empty:
                        logging.info(f"API: Successfully read Excel file with {len(df)} rows using engine: {engine}")
                        break
                        
                except UnicodeDecodeError as ude:
                    logging.warning(f"API: Unicode decode error with engine {engine}: {str(ude)}")
                    continue
                except Exception as ee:
                    logging.warning(f"API: Error reading with engine {engine}: {str(ee)}")
                    continue
        else:
            # Handle CSV files
            df = pd.read_csv(file_path, encoding='utf-8')
        
        if df is None or df.empty:
            return jsonify({'error': 'Не вдалося прочитати файл. Перевірте кодування або формат.'}), 500
        
        total_lines = len(df)
        success_count = 0
        error_count = 0
        sample_data = []
        
        # Debug: показати назви колонок
        logging.info(f"API: Колонки в файлі: {list(df.columns)}")
        
        # Process DataFrame instead of CSV reader
        for i, (index, row) in enumerate(df.iterrows()):
            # Знайти колонку з ЄДРПОУ (гнучкий пошук)
            edrpou = ''
            for col_name in df.columns:
                if any(keyword in col_name.upper() for keyword in ['ЄДРПОУ', 'EDRPOU']):
                    if not pd.isna(row[col_name]):
                        edrpou = str(row[col_name]).strip()
                        break
            
            # Debug: показати кілька прикладів ЄДРПОУ
            if i < 5:
                logging.info(f"API: Рядок {i}: ЄДРПОУ = '{edrpou}'")
            
            if edrpou and edrpou != 'nan' and len(edrpou) >= 6 and edrpou.isdigit():
                success_count += 1
                if len(sample_data) < 5:
                    sample_data.append({
                        'edrpou': edrpou,
                        'name': str(row.get('Название компании', '') or row.get('Назва', '') or '')[:50],
                        'kved': str(row.get('КВЕД', '') or row.get('KVED', '') or ''),
                        'region': str(row.get('Область', '') or row.get('Region', '') or ''),
                        'revenue': str(row.get('Чистий дохід від реалізації продукції    ', '') or row.get('Revenue', '') or ''),
                        'profit': str(row.get('Чистий фінансовий результат: прибуток                                           ', '') or row.get('Profit', '') or '')
                    })
            else:
                error_count += 1
        
        return jsonify({
            'success': True,
            'total_lines': total_lines,
            'processed': total_lines,
            'success_count': success_count,
            'error_count': error_count,
            'sample_data': sample_data,
            'message': f'Файл проаналізовано повністю. Всього рядків: {total_lines}. Валідних ЄДРПОУ: {success_count}, помилкових: {error_count}. Для повної обробки в БД потрібен окремий процес.'
        })
        
    except Exception as e:
        logging.error(f"Error analyzing CSV: {e}")
        return jsonify({'error': f'Помилка аналізу файлу: {str(e)}'}), 500

@main.route('/database-import/<filename>')
@login_required
@admin_required
def database_import(filename):
    """Show database import progress page"""
    
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, filename)
    
    if not os.path.exists(file_path) or not filename.endswith('_processed.csv'):
        flash(f'Файл {filename} не знайдено або це не оброблений CSV файл.', 'error')
        return redirect(url_for('main.file_manager'))
    
    return render_template('database_import_progress.html', filename=filename)

# Global variable to track import status
import_status = {}

@main.route('/api/start-database-import/<filename>', methods=['POST'])
@login_required
@admin_required
def start_database_import(filename):
    """Start database import in background"""
    
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, filename)
    
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    try:
        # Initialize status
        import_status[filename] = {
            'running': True,
            'processed': 0,
            'total': 0,
            'updated': 0,
            'inserted': 0,
            'errors': 0,
            'success': False,
            'message': 'Підготовка...'
        }
        
        # Start background process
        import threading
        thread = threading.Thread(target=run_database_import, args=(filename, file_path))
        thread.daemon = True
        thread.start()
        
        return jsonify({'success': True, 'message': 'Import started'})
        
    except Exception as e:
        logging.error(f"Error starting import: {e}")
        return jsonify({'error': f'Помилка запуску: {str(e)}'}), 500

@main.route('/api/database-import-status/<filename>')
@login_required
def database_import_status(filename):
    """Get database import status"""
    status = import_status.get(filename, {
        'running': False,
        'processed': 0,
        'total': 0,
        'updated': 0,
        'inserted': 0,
        'errors': 0,
        'success': False,
        'message': 'Не знайдено'
    })
    
    return jsonify(status)

def run_database_import(filename, file_path):
    """Run database import in background thread with optimizations"""
    from app import app, db
    import csv
    
    with app.app_context():
        try:
            # Оптимізація PostgreSQL для швидшого завантаження (тільки дозволені параметри)
            try:
                db.session.execute(db.text("SET synchronous_commit = off;"))
                db.session.commit()
            except Exception:
                # Ігноруємо помилки з недоступними параметрами в production
                pass
            
            # Count total lines
            with open(file_path, 'r', encoding='utf-8') as f:
                total_lines = sum(1 for _ in f) - 1  # exclude header
            
            import_status[filename]['total'] = total_lines
            import_status[filename]['message'] = f'Початок обробки {total_lines} записів'
            
            success_count = 0
            update_count = 0
            insert_count = 0
            error_count = 0
            
            # Process CSV file
            with open(file_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                
                for row_num, row in enumerate(reader, 1):
                    try:
                        edrpou = row.get('edrpou', '').strip()
                        if not edrpou:
                            error_count += 1
                            continue
                        
                        # Check if company exists (optimized query)
                        existing = db.session.execute(
                            db.text("SELECT 1 FROM companies WHERE edrpou = :edrpou LIMIT 1"), 
                            {'edrpou': edrpou}
                        ).fetchone()
                        
                        if existing:
                            # Update existing company
                            db.session.execute(db.text("""
                                UPDATE companies SET 
                                    name = COALESCE(:name, name),
                                    kved_code = COALESCE(:kved_code, kved_code),
                                    kved_description = COALESCE(:kved_description, kved_description),
                                    region_name = COALESCE(:region_name, region_name),
                                    phone = COALESCE(:phone, phone),
                                    address = COALESCE(:address, address),
                                    company_size_name = COALESCE(:company_size_name, company_size_name),
                                    personnel_2019 = COALESCE(:personnel_2019, personnel_2019),
                                    revenue_2019 = COALESCE(:revenue_2019, revenue_2019),
                                    profit_2019 = COALESCE(:profit_2019, profit_2019),
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE edrpou = :edrpou
                            """), {
                                'name': row.get('name', '')[:500] if row.get('name') else None,
                                'kved_code': row.get('kved_code', '')[:20] if row.get('kved_code') else None,
                                'kved_description': row.get('kved_description', '')[:500] if row.get('kved_description') else None,
                                'region_name': row.get('region_name', '')[:100] if row.get('region_name') else None,
                                'phone': row.get('phone', '')[:50] if row.get('phone') else None,
                                'address': row.get('address', '')[:500] if row.get('address') else None,
                                'company_size_name': row.get('company_size_name', '')[:50] if row.get('company_size_name') else None,
                                'personnel_2019': int(float(row.get('personnel_2019', 0))) if row.get('personnel_2019') else None,
                                'revenue_2019': float(row.get('revenue_2019', 0)) if row.get('revenue_2019') else None,
                                'profit_2019': float(row.get('profit_2019', 0)) if row.get('profit_2019') else None,
                                'edrpou': edrpou
                            })
                            update_count += 1
                        else:
                            # Insert new company
                            db.session.execute(db.text("""
                                INSERT INTO companies (
                                    edrpou, name, kved_code, kved_description, region_name,
                                    phone, address, company_size_name, personnel_2019,
                                    revenue_2019, profit_2019, source, actualized, created_at
                                ) VALUES (:edrpou, :name, :kved_code, :kved_description, :region_name, 
                                         :phone, :address, :company_size_name, :personnel_2019,
                                         :revenue_2019, :profit_2019, :source, :actualized, CURRENT_TIMESTAMP)
                            """), {
                                'edrpou': edrpou,
                                'name': row.get('name', '')[:500] if row.get('name') else None,
                                'kved_code': row.get('kved_code', '')[:20] if row.get('kved_code') else None,
                                'kved_description': row.get('kved_description', '')[:500] if row.get('kved_description') else None,
                                'region_name': row.get('region_name', '')[:100] if row.get('region_name') else None,
                                'phone': row.get('phone', '')[:50] if row.get('phone') else None,
                                'address': row.get('address', '')[:500] if row.get('address') else None,
                                'company_size_name': row.get('company_size_name', '')[:50] if row.get('company_size_name') else None,
                                'personnel_2019': int(float(row.get('personnel_2019', 0))) if row.get('personnel_2019') else None,
                                'revenue_2019': float(row.get('revenue_2019', 0)) if row.get('revenue_2019') else None,
                                'profit_2019': float(row.get('profit_2019', 0)) if row.get('profit_2019') else None,
                                'source': 'імпорт',
                                'actualized': 'так'
                            })
                            insert_count += 1
                        
                        success_count += 1
                        
                        # Update status every 500 records and commit in batches
                        if row_num % 500 == 0:
                            db.session.commit()
                            import_status[filename].update({
                                'processed': row_num,
                                'updated': update_count,
                                'inserted': insert_count,
                                'errors': error_count,
                                'message': f'Оброблено {row_num}/{total_lines}'
                            })
                    
                    except Exception as e:
                        error_count += 1
                        logging.error(f"Error processing row {row_num}: {e}")
                        continue
            
            # Final commit
            db.session.commit()
            
            # Повернути стандартні налаштування PostgreSQL
            try:
                db.session.execute(db.text("SET synchronous_commit = on;"))
                db.session.commit()
            except Exception:
                # Ігноруємо помилки з недоступними параметрами в production
                pass
            
            # Update final status
            import_status[filename].update({
                'running': False,
                'processed': success_count,
                'updated': update_count,
                'inserted': insert_count,
                'errors': error_count,
                'success': True,
                'message': f'Завершено: {success_count} записів'
            })
            
        except Exception as e:
            logging.error(f"Critical error in database import: {e}")
            import_status[filename].update({
                'running': False,
                'success': False,
                'error': str(e),
                'message': f'Критична помилка: {e}'
            })

def process_csv_to_database(file_path):
    """Process CSV file to database with proper field mapping"""
    success_count = 0
    error_count = 0
    
    try:
        with open(file_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            
            batch_data = []
            row_count = 0
            
            for row in reader:
                row_count += 1
                try:
                    # Map CSV columns to database fields
                    edrpou = str(row.get('Код ЄДРПОУ', '') or row.get('ЄДРПОУ', '')).strip()
                    if not edrpou or edrpou == 'nan':
                        error_count += 1
                        continue
                    
                    # Extract all fields from CSV
                    company_data = {
                        'edrpou': edrpou,
                        'name': (row.get('Название компании', '') or row.get('Назва', '') or row.get('Назва компанії', ''))[:500],
                        'kved_code': str(row.get('КВЕД', '') or '').strip()[:20],
                        'kved_description': (row.get('Основний вид діяльності (КВЕД)', '') or row.get('КВЕД опис', ''))[:500],
                        'region_name': (row.get('Область', '') or row.get('Регіон', ''))[:100],
                        'phone': str(row.get('Tелефон', '') or row.get('Телефон', ''))[:50],
                        'address': (row.get('Адреса реєстрації', '') or row.get('Адреса', ''))[:500],
                        'company_size_name': (row.get('Размер', '') or row.get('Розмір компанії', ''))[:50]
                    }
                    
                    # Parse numeric fields safely
                    try:
                        personnel_str = str(row.get('Персонал (2019 р.)', '') or row.get('Кількість працівників', '') or row.get('Персонал (осіб)', ''))
                        if personnel_str and personnel_str.replace('.', '').replace(',', '').isdigit():
                            company_data['personnel_2019'] = int(float(personnel_str))
                    except:
                        company_data['personnel_2019'] = None
                    
                    try:
                        revenue_str = str(row.get('Чистий дохід від реалізації продукції    ', '') or row.get('Дохід', '') or row.get('Чистий дохід (тис. грн)', ''))
                        if revenue_str and revenue_str.replace(',', '.').replace(' ', ''):
                            company_data['revenue_2019'] = float(revenue_str.replace(',', '.'))
                    except:
                        company_data['revenue_2019'] = None
                    
                    try:
                        profit_str = str(row.get('Чистий фінансовий результат: прибуток                                           ', '') or row.get('Прибуток', '') or row.get('Прибуток (збиток)', ''))
                        if profit_str and profit_str.replace(',', '.').replace(' ', ''):
                            company_data['profit_2019'] = float(profit_str.replace(',', '.'))
                    except:
                        company_data['profit_2019'] = None
                    
                    batch_data.append(company_data)
                    
                    # Process in batches of 100
                    if len(batch_data) >= 100:
                        batch_success = process_csv_batch_to_db(batch_data)
                        success_count += batch_success
                        error_count += (len(batch_data) - batch_success)
                        batch_data = []
                        
                        if row_count % 1000 == 0:
                            logging.info(f"Processed {row_count} rows: {success_count} success, {error_count} errors")
                            
                except Exception as e:
                    logging.error(f"Error processing row {row_count}: {e}")
                    error_count += 1
                    continue
            
            # Process remaining batch
            if batch_data:
                batch_success = process_csv_batch_to_db(batch_data)
                success_count += batch_success
                error_count += (len(batch_data) - batch_success)
    
    except Exception as e:
        logging.error(f"Error processing CSV file: {e}")
        error_count += 100
    
    return success_count, error_count

def process_csv_batch_to_db(companies_data):
    """Process batch of companies with all fields to database"""
    success_count = 0
    
    try:
        # Get existing companies for this batch
        edrpou_list = [data['edrpou'] for data in companies_data]
        existing_companies = {
            company.edrpou: company for company in 
            db.session.query(Company).filter(Company.edrpou.in_(edrpou_list)).all()
        }
        
        for data in companies_data:
            try:
                edrpou = data['edrpou']
                
                if edrpou in existing_companies:
                    # Update existing company with ALL fields
                    company = existing_companies[edrpou]
                    company.name = data['name'] if data['name'] else company.name
                    company.kved_code = data['kved_code'] if data['kved_code'] else company.kved_code
                    company.kved_description = data['kved_description'] if data['kved_description'] else company.kved_description
                    company.region_name = data['region_name'] if data['region_name'] else company.region_name
                    company.phone = data['phone'] if data['phone'] else company.phone
                    company.address = data['address'] if data['address'] else company.address
                    company.company_size_name = data['company_size_name'] if data['company_size_name'] else company.company_size_name
                    
                    if data['personnel_2019'] is not None:
                        company.personnel_2019 = data['personnel_2019']
                    if data['revenue_2019'] is not None:
                        company.revenue_2019 = data['revenue_2019']
                    if data['profit_2019'] is not None:
                        company.profit_2019 = data['profit_2019']
                else:
                    # Create new company with ALL fields
                    company = Company(
                        edrpou=edrpou,
                        name=data['name'] if data['name'] else None,
                        kved_code=data['kved_code'] if data['kved_code'] else None,
                        kved_description=data['kved_description'] if data['kved_description'] else None,
                        region_name=data['region_name'] if data['region_name'] else None,
                        phone=data['phone'] if data['phone'] else None,
                        address=data['address'] if data['address'] else None,
                        company_size_name=data['company_size_name'] if data['company_size_name'] else None,
                        personnel_2019=data['personnel_2019'],
                        revenue_2019=data['revenue_2019'],
                        profit_2019=data['profit_2019'],
                        source='основний',
                        actualized='ні'
                    )
                    db.session.add(company)
                
                success_count += 1
                
            except Exception as e:
                logging.error(f"Error processing company {data['edrpou']}: {e}")
                continue
        
        # Single commit for entire batch
        db.session.commit()
        
    except Exception as e:
        logging.error(f"Error in CSV batch processing: {e}")
        db.session.rollback()
        success_count = 0
    
    return success_count

@main.route('/companies')
@login_required
def companies():
    if not current_user.has_permission('view'):
        flash('У вас немає прав для перегляду компаній.', 'danger')
        return redirect(url_for('main.index'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Changed to 20 per page as requested
    
    # Search functionality
    search_edrpou = request.args.get('search_edrpou', type=str, default='').strip()
    
    # Build query with optional filters
    query = db.select(Company)
    
    # EDRPOU search
    if search_edrpou:
        query = query.where(Company.edrpou.ilike(f'%{search_edrpou}%'))
    
    # Filter by region name
    region_name = request.args.get('region_name', type=str)
    if region_name:
        query = query.where(Company.region_name == region_name)
    
    # Filter by KVED code
    kved_code = request.args.get('kved_code', type=str)
    if kved_code and not search_edrpou:  # Don't apply filter if search is active
        query = query.where(Company.kved_code == kved_code)
    
    # Filter by company size
    size_id = request.args.get('size_id', type=str)
    if size_id and not search_edrpou:  # Don't apply filter if search is active
        query = query.where(Company.company_size_name == size_id)
    
    # Sorting
    sort_by = request.args.get('sort_by', 'name')
    sort_order = request.args.get('sort_order', 'asc')
    
    if sort_by == 'revenue':
        query = query.order_by(desc(Company.revenue_2019) if sort_order == 'desc' else asc(Company.revenue_2019))
    elif sort_by == 'profit':
        query = query.order_by(desc(Company.profit_2019) if sort_order == 'desc' else asc(Company.profit_2019))
    elif sort_by == 'personnel':
        query = query.order_by(desc(Company.personnel_2019) if sort_order == 'desc' else asc(Company.personnel_2019))
    elif sort_by == 'ranking':
        query = query.order_by(asc(Company.ranking) if sort_order == 'asc' else desc(Company.ranking))
    else:
        query = query.order_by(asc(Company.name) if sort_order == 'asc' else desc(Company.name))
    
    # Get total count for pagination
    total_companies = len(db.session.execute(query).scalars().all())
    
    # Apply pagination
    query = query.offset((page - 1) * per_page).limit(per_page)
    companies = db.session.execute(query).scalars().all()
    
    # Get unique values for filters (filter out None values)
    regions = db.session.execute(
        db.select(Company.region_name).distinct().where(Company.region_name.isnot(None))
    ).scalars().all()
    kveds = db.session.execute(
        db.select(Company.kved_code).distinct().where(Company.kved_code.isnot(None))
    ).scalars().all()
    sizes = db.session.execute(
        db.select(Company.company_size_name).distinct().where(Company.company_size_name.isnot(None))
    ).scalars().all()
    
    # Create proper pagination object
    class Pagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None
            
        def iter_pages(self, left_edge=2, left_current=2, right_current=3, right_edge=2):
            last = self.pages
            for num in range(1, last + 1):
                if num <= left_edge or \
                   (num > self.page - left_current - 1 and num < self.page + right_current) or \
                   num > last - right_edge:
                    yield num
    
    companies_paginated = Pagination(companies, page, per_page, total_companies)
    
    return render_template('companies.html', 
                         companies=companies_paginated,
                         regions=regions, 
                         kveds=kveds, 
                         sizes=sizes,
                         current_filters={
                             'region_name': region_name,
                             'kved_code': kved_code,
                             'size_id': size_id,
                             'sort_by': sort_by,
                             'sort_order': sort_order,
                             'search_edrpou': search_edrpou
                         })

@main.route('/filter', methods=['GET', 'POST'])
@login_required
def filter_companies_route():
    if not current_user.has_permission('edit'):
        flash('У вас немає прав для фільтрації компаній.', 'danger')
        return redirect(url_for('main.index'))
    
    if request.method == 'POST':
        # Get filter parameters from form
        min_employees = int(request.form.get('min_employees', 0) or 0)
        min_revenue = float(request.form.get('min_revenue', 0) or 0)
        min_profit = request.form.get('min_profit')
        min_profit = float(min_profit) if min_profit and min_profit.strip() else None
        
        selected_regions = request.form.getlist('region_filter')
        selected_kved = request.form.getlist('kved_filter') 
        selected_sizes = request.form.getlist('size_filter')
        
        try:
            # Start with all companies
            from sqlalchemy import select, and_
            query = select(Company)
            
            # Apply basic filters (Stage 1)
            filters = []
            filter_description = []
            
            if min_employees > 0:
                filters.append(Company.personnel_2019 >= min_employees)
                filter_description.append(f"Employees >= {min_employees}")
            if min_revenue > 0:
                filters.append(Company.revenue_2019 >= min_revenue)
                filter_description.append(f"Revenue >= {min_revenue:,.0f}")
            if min_profit is not None:
                filters.append(Company.profit_2019 >= min_profit)
                filter_description.append(f"Profit >= {min_profit:,.0f}")
                
            # Apply additional filters (Stage 2)
            if selected_regions:
                filters.append(Company.region_name.in_(selected_regions))
                filter_description.append(f"Regions: {', '.join(selected_regions[:3])}{'...' if len(selected_regions) > 3 else ''}")
            if selected_kved:
                filters.append(Company.kved_code.in_(selected_kved))
                filter_description.append(f"KVED: {', '.join(selected_kved)}")
            if selected_sizes:
                filters.append(Company.company_size_name.in_(selected_sizes))
                filter_description.append(f"Sizes: {', '.join(selected_sizes)}")
                
            # Combine all filters
            if filters:
                query = query.where(and_(*filters))
                
            # Use raw SQL to avoid SQLAlchemy UTF-8 issues
            try:
                # Build SQL query manually 
                sql_conditions = []
                sql_params = {}
                
                if min_employees > 0:
                    sql_conditions.append("personnel_2019 >= :min_employees")
                    sql_params['min_employees'] = min_employees
                if min_revenue > 0:
                    sql_conditions.append("revenue_2019 >= :min_revenue")
                    sql_params['min_revenue'] = min_revenue
                if min_profit is not None:
                    sql_conditions.append("profit_2019 >= :min_profit")
                    sql_params['min_profit'] = min_profit
                
                # Build WHERE clause
                where_clause = " AND ".join(sql_conditions) if sql_conditions else "1=1"
                
                # Count query only (safer than loading full objects)
                count_sql = f"SELECT COUNT(*) FROM companies WHERE {where_clause}"
                count_result = db.session.execute(db.text(count_sql), sql_params)
                count = count_result.scalar() or 0
                
                # Don't load actual company objects to avoid UTF-8 issues
                selected_companies = []
                
                logging.info(f"Raw SQL count query successful: {count} companies found")
                
            except Exception as e:
                logging.error(f"Error in SQL query: {e}")
                count = 0
                selected_companies = []
                
            # Save selection criteria for display
            criteria_text = " | ".join(filter_description) if filter_description else "No filters"
            
            # Create SelectionBase with minimal operations (avoid UTF-8 issues)
            try:
                # Include selection criteria in SelectionBase
                selection_base = SelectionBase(
                    name=f'Selection Base ({count} companies)',
                    companies_count=count,
                    min_employees=min_employees if min_employees > 0 else None,
                    min_revenue=min_revenue if min_revenue > 0 else None,
                    min_profit=min_profit if min_profit is not None else None,
                    is_active=True
                )
                
                # Add and commit immediately
                db.session.add(selection_base)
                db.session.commit()
                logging.info(f"Created SelectionBase ID: {selection_base.id} with {count} companies")
                
                # Skip adding SelectionCompany records for now to avoid UTF-8 issues
                # We'll store company IDs in session instead
                logging.info(f"Skipping SelectionCompany creation - will use session storage")
                
            except Exception as e:
                db.session.rollback()
                logging.error(f"Error creating SelectionBase: {e}")
                flash('Error creating selection base.', 'danger')
                return redirect(request.url)
            
            session['selection_criteria'] = criteria_text
            # Store empty for now since we don't load company objects (avoiding UTF-8 issues)
            session['selected_company_ids'] = []
            session['selection_count'] = count
            
            # Clear previous ranking info from session
            session.pop('last_ranking_name', None)
            session.pop('last_ranking_count', None) 
            session.pop('last_ranking_criteria', None)
            
            # Save selection info for display in ranking page
            selection_info = {
                'total_companies': count,
                'min_personnel': min_employees if min_employees > 0 else 'Not set',
                'min_revenue': f"{min_revenue:,.0f}" if min_revenue > 0 else 'Not set',
                'min_profit': f"{min_profit:,.0f}" if min_profit is not None else 'Not set',
                'regions_count': len(selected_regions) if selected_regions else 'All regions',
                'kved_count': len(selected_kved) if selected_kved else 'All KVED codes'
            }
            session['selection_info'] = selection_info
            
            flash(f'Created selection base: {count} companies with criteria "{criteria_text}"', 'success')
            return redirect(url_for('main.export'))  # Go to ranking stage
            
        except Exception as e:
            logging.error(f"Error in selection: {e}")
            flash('Error creating selection base.', 'danger')
            return redirect(request.url)
    
    # Load filter options for GET request - using simplified structure
    regions = db.session.execute(
        db.select(Company.region_name).distinct().where(Company.region_name.isnot(None))
    ).scalars().all()
    kveds = db.session.execute(
        db.select(Company.kved_code).distinct().where(Company.kved_code.isnot(None))
    ).scalars().all()
    sizes = db.session.execute(
        db.select(Company.company_size_name).distinct().where(Company.company_size_name.isnot(None))
    ).scalars().all()
    
    return render_template('filter.html', regions=regions, kveds=kveds, sizes=sizes)

@main.route('/api/selection-history')
@login_required
def selection_history():
    """Get selection history for filter page"""
    try:
        selection_bases = db.session.execute(
            db.select(SelectionBase).order_by(SelectionBase.created_at.desc()).limit(10)
        ).scalars().all()
        
        history = []
        for sb in selection_bases:
            criteria_parts = []
            if sb.min_employees:
                criteria_parts.append(f"Працівників ≥ {sb.min_employees}")
            if sb.min_revenue:
                criteria_parts.append(f"Дохід ≥ {sb.min_revenue:,.0f}")
            if sb.min_profit:
                criteria_parts.append(f"Прибуток ≥ {sb.min_profit:,.0f}")
            
            history.append({
                'id': sb.id,
                'companies_count': sb.companies_count,
                'criteria_display': " | ".join(criteria_parts) if criteria_parts else "Без критеріїв",
                'created_at': sb.created_at.isoformat() if sb.created_at else None,
                'min_employees': sb.min_employees,
                'min_revenue': float(sb.min_revenue) if sb.min_revenue else None,
                'min_profit': float(sb.min_profit) if sb.min_profit else None
            })
        
        return jsonify({'success': True, 'history': history})
        
    except Exception as e:
        logging.error(f"Error getting selection history: {e}")
        return jsonify({'success': False, 'error': str(e)})

@main.route('/api/selection-history/<int:selection_id>/apply', methods=['POST'])
@login_required
def apply_selection_history(selection_id):
    """Apply historical selection criteria"""
    try:
        selection_base = db.session.execute(
            db.select(SelectionBase).where(SelectionBase.id == selection_id)
        ).scalar_one_or_none()
        
        if not selection_base:
            return jsonify({'success': False, 'error': 'Відбір не знайдено'})
        
        return jsonify({
            'success': True,
            'selection': {
                'id': selection_base.id,
                'min_employees': selection_base.min_employees,
                'min_revenue': float(selection_base.min_revenue) if selection_base.min_revenue else None,
                'min_profit': float(selection_base.min_profit) if selection_base.min_profit else None,
                'created_at': selection_base.created_at.isoformat() if selection_base.created_at else None
            }
        })
        
    except Exception as e:
        logging.error(f"Error applying selection history: {e}")
        return jsonify({'success': False, 'error': str(e)})

@main.route('/api/company/<int:company_id>')
@login_required
def get_company_details(company_id):
    if not current_user.has_permission('view'):
        return {'success': False, 'error': 'Немає прав доступу'}, 403
    
    try:
        company = db.session.execute(
            db.select(Company).where(Company.id == company_id)
        ).scalar_one_or_none()
        
        if not company:
            return {'success': False, 'error': 'Компанію не знайдено'}, 404
        
        # Get ranking history for this company (safe version)
        ranking_history = []
        try:
            from sqlalchemy import text
            # Спробуємо отримати історію рейтингів, якщо таблиця існує
            ranking_history_result = db.session.execute(text(
                "SELECT ranking_name, ranking_position, ranking_criteria, source_name, created_at FROM company_ranking_history WHERE company_id = :company_id ORDER BY created_at DESC"
            ), {'company_id': company.id})
            
            for row in ranking_history_result:
                ranking_history.append({
                    'ranking_name': row[0],
                    'ranking_position': row[1], 
                    'ranking_criteria': row[2],
                    'source_name': row[3],
                    'created_at': row[4].isoformat() if row[4] else None
                })
        except Exception as e:
            # Якщо таблиця неповна або відсутні колонки - просто ігноруємо
            ranking_history = []
        
        # Convert to dict with all fields
        company_data = {
            'id': company.id,
            'edrpou': company.edrpou,
            'name': company.name,
            'kved_code': company.kved_code,
            'kved_description': company.kved_description,
            'personnel_2019': company.personnel_2019,
            'region_name': company.region_name,
            'phone': company.phone,
            'address': company.address,
            'revenue_2019': float(company.revenue_2019) if company.revenue_2019 else None,
            'profit_2019': float(company.profit_2019) if company.profit_2019 else None,
            'company_size_name': company.company_size_name,
            'first_name': company.first_name,
            'middle_name': company.middle_name,
            'last_name': company.last_name,
            'work_phone': company.work_phone,
            'corporate_site': company.corporate_site,
            'work_email': company.work_email,
            'company_status': company.company_status,
            'director': company.director,
            'government_purchases': float(company.government_purchases) if company.government_purchases else None,
            'tender_count': company.tender_count,
            'initials': company.initials,
            'source': company.source,
            'actualized': company.actualized,
            'ranking': company.ranking,
            'ranking_criteria': getattr(company, 'ranking_criteria', None),
            'created_at': company.created_at.isoformat() if company.created_at else None,
            'updated_at': company.updated_at.isoformat() if company.updated_at else None,
            'ranking_history': ranking_history
        }
        
        return {'success': True, 'company': company_data}
        
    except Exception as e:
        logging.error(f"Error getting company details: {str(e)}")
        return {'success': False, 'error': str(e)}, 500

@main.route('/rankings')
@login_required
def rankings_list():
    """Сторінка зі списком всіх рейтингів"""
    if not current_user.has_permission('view'):
        flash('У вас немає прав для перегляду рейтингів.', 'danger')
        return redirect(url_for('main.index'))
        
    try:
        # Отримуємо всі рейтинги
        rankings = db.session.execute(
            db.select(Ranking).order_by(Ranking.created_at.desc())
        ).scalars().all()
        
        rankings_data = []
        for ranking in rankings:
            # Окремо отримуємо SelectionBase для кожного рейтингу
            selection_base = db.session.execute(
                db.select(SelectionBase).where(SelectionBase.id == ranking.selection_base_id)
            ).scalar_one_or_none()
            # Отримуємо першу компанію з рейтингу для превью (спрощено)
            first_ranking_company = db.session.execute(
                db.select(RankingCompany)
                .where(RankingCompany.ranking_id == ranking.id)
                .order_by(RankingCompany.position)
                .limit(1)
            ).scalar_one_or_none()
            
            if first_ranking_company:
                first_company = db.session.execute(
                    db.select(Company).where(Company.id == first_ranking_company.company_id)
                ).scalar_one_or_none()
                first_company_name = first_company.name if first_company else "Немає компаній"
            else:
                first_company_name = "Немає компаній"
            
            rankings_data.append({
                'ranking': ranking,
                'selection_base': selection_base,
                'first_company_name': first_company_name
            })
        
        return render_template('rankings_list.html', rankings=rankings_data)
        
    except Exception as e:
        logging.error(f"Error loading rankings list: {e}")
        flash('Помилка завантаження списку рейтингів', 'danger')
        return redirect(url_for('main.index'))

@main.route('/ranking/<int:ranking_id>/view')
@login_required
def view_ranking(ranking_id):
    """Перегляд конкретного рейтингу з компаніями"""
    if not current_user.has_permission('view'):
        flash('У вас немає прав для перегляду рейтингів.', 'danger')
        return redirect(url_for('main.index'))
        
    try:
        # Отримуємо рейтинг
        ranking = db.session.execute(
            db.select(Ranking).where(Ranking.id == ranking_id)
        ).scalar_one_or_none()
        
        if not ranking:
            flash('Рейтинг не знайдено', 'error')
            return redirect(url_for('main.rankings_list'))
        
        # Отримуємо компанії рейтингу (спрощено без JOIN)
        ranking_companies_data = []
        ranking_companies = db.session.execute(
            db.select(RankingCompany)
            .where(RankingCompany.ranking_id == ranking_id)
            .order_by(RankingCompany.position)
        ).scalars().all()
        
        for ranking_company in ranking_companies:
            company = db.session.execute(
                db.select(Company).where(Company.id == ranking_company.company_id)
            ).scalar_one_or_none()
            
            if company:
                ranking_companies_data.append((ranking_company, company))
        
        # Отримуємо базу відбору
        selection_base = db.session.execute(
            db.select(SelectionBase).where(SelectionBase.id == ranking.selection_base_id)
        ).scalar_one_or_none()
        
        return render_template('ranking_view.html', 
                             ranking=ranking, 
                             companies=ranking_companies_data,
                             selection_base=selection_base)
        
    except Exception as e:
        logging.error(f"Error loading ranking {ranking_id}: {e}")
        flash('Помилка завантаження рейтингу', 'danger')
        return redirect(url_for('main.rankings_list'))

@main.route('/ranking', methods=['GET', 'POST'])
@login_required
def ranking():
    """New ranking page with clean implementation"""
    if not current_user.has_permission('view'):
        flash('У вас немає прав для створення рейтингу.', 'danger')
        return redirect(url_for('main.index'))
    
    if request.method == 'POST':
        try:
            # Get form data as JSON with better error handling
            logging.info(f"Request content type: {request.content_type}")
            logging.info(f"Request is_json: {request.is_json}")
            
            if request.is_json:
                try:
                    data = request.get_json(force=True)
                    logging.info(f"Successfully parsed JSON data: {data}")
                except Exception as json_error:
                    logging.error(f"JSON parsing error: {json_error}")
                    logging.info(f"Raw request data: {request.get_data()}")
                    return jsonify({'success': False, 'error': 'Некоректний JSON запит'})
            else:
                # Convert form data to expected format
                data = {
                    'kved_filter': request.form.getlist('kved_filter'),
                    'region_filter': request.form.getlist('region_filter'),
                    'size_filter': request.form.getlist('size_filter'),
                    'sort_criteria': request.form.get('sort_criteria'),
                    'sort_order': request.form.get('sort_order'),
                    'year_source': request.form.get('year_source'),
                    'ranking_name': request.form.get('ranking_name')
                }
                logging.info(f"Form data converted to: {data}")
            
            # Validate required fields
            if not data.get('sort_criteria') or not data.get('ranking_name'):
                return jsonify({'success': False, 'error': 'Заповніть всі обов\'язкові поля'})
            
            # Get latest active SelectionBase
            selection_base = db.session.execute(
                db.select(SelectionBase)
                .where(SelectionBase.is_active == True)
                .order_by(SelectionBase.created_at.desc())
            ).first()
            
            if not selection_base:
                return jsonify({'success': False, 'error': 'Спочатку створіть базу відбору на сторінці фільтрації компаній'})
            
            selection_base = selection_base[0] if selection_base else None
            logging.info(f"Using SelectionBase ID: {selection_base.id if selection_base else 'None'}")
            
            # Use raw SQL to recreate the base selection + additional filters
            try:
                # Build SQL conditions for base selection (from SelectionBase)
                base_conditions = []
                base_params = {}
                
                if selection_base.min_employees:
                    base_conditions.append("personnel_2019 >= :min_employees")
                    base_params['min_employees'] = selection_base.min_employees
                if selection_base.min_revenue:
                    base_conditions.append("revenue_2019 >= :min_revenue")
                    base_params['min_revenue'] = selection_base.min_revenue
                if selection_base.min_profit:
                    base_conditions.append("profit_2019 >= :min_profit")
                    base_params['min_profit'] = selection_base.min_profit
                
                # Add additional ranking filters
                if data['kved_filter'] and data['kved_filter'] != ['']:
                    kved_placeholders = ', '.join([f":kved_{i}" for i in range(len(data['kved_filter']))])
                    base_conditions.append(f"kved_code IN ({kved_placeholders})")
                    for i, kved in enumerate(data['kved_filter']):
                        base_params[f'kved_{i}'] = kved
                
                if data['region_filter'] and data['region_filter'] != ['']:
                    region_placeholders = ', '.join([f":region_{i}" for i in range(len(data['region_filter']))])
                    base_conditions.append(f"region_name IN ({region_placeholders})")
                    for i, region in enumerate(data['region_filter']):
                        base_params[f'region_{i}'] = region
                
                # Build final query
                where_clause = " AND ".join(base_conditions) if base_conditions else "1=1"
                
                # Log detailed information about the query
                logging.info(f"SelectionBase criteria: min_employees={selection_base.min_employees}, min_revenue={selection_base.min_revenue}, min_profit={selection_base.min_profit}")
                logging.info(f"Additional filters: kved={data.get('kved_filter')}, region={data.get('region_filter')}")
                logging.info(f"SQL WHERE clause: {where_clause}")
                logging.info(f"SQL parameters: {base_params}")
                
                # Use raw SQL to get all needed company fields
                companies_sql = f"""SELECT id, edrpou, name, revenue_2019, profit_2019, personnel_2019, 
                    kved_code, kved_description, region_name, phone, address,
                    company_size_name, first_name, middle_name, last_name, work_phone,
                    corporate_site, work_email, company_status, director, 
                    government_purchases, tender_count, initials, actualized
                    FROM companies WHERE {where_clause} ORDER BY revenue_2019 DESC NULLS LAST LIMIT 1000"""
                
                logging.info(f"Final SQL query: {companies_sql}")
                
                result = db.session.execute(db.text(companies_sql), base_params)
                companies_raw = result.fetchall()
                
                logging.info(f"Raw SQL found {len(companies_raw)} companies for ranking")
                
                # Convert to objects for sorting
                companies = []
                for row in companies_raw:
                    company_dict = {
                        'id': row[0],
                        'edrpou': row[1] or '',
                        'name': row[2] or '',
                        'revenue_2019': float(row[3]) if row[3] else 0,
                        'profit_2019': float(row[4]) if row[4] else 0,
                        'personnel_2019': int(row[5]) if row[5] else 0,
                        'kved_code': row[6] or '',
                        'kved_description': row[7] or '',
                        'region_name': row[8] or '',
                        'phone': row[9] or '',
                        'address': row[10] or '',
                        'company_size_name': row[11] or '',
                        'first_name': row[12] or '',
                        'middle_name': row[13] or '',
                        'last_name': row[14] or '',
                        'work_phone': row[15] or '',
                        'corporate_site': row[16] or '',
                        'work_email': row[17] or '',
                        'company_status': row[18] or '',
                        'director': row[19] or '',
                        'government_purchases': float(row[20]) if row[20] else 0,
                        'tender_count': int(row[21]) if row[21] else 0,
                        'initials': row[22] or '',
                        'actualized': row[23] or ''
                    }
                    companies.append(company_dict)
                    
            except Exception as e:
                logging.error(f"Error in raw SQL companies query: {e}")
                return jsonify({'success': False, 'error': f'Помилка при отриманні компаній: {str(e)}'})
            
            logging.info(f"Found {len(companies)} companies from SelectionBase after applying filters")
            
            # Sort by criteria
            sort_criteria = data['sort_criteria']
            sort_order = data.get('sort_order', 'desc')
            reverse_sort = sort_order == 'desc'
            
            # Sort companies (now using dictionary objects)
            if sort_criteria == 'revenue':
                sorted_companies = sorted(companies, key=lambda x: x['revenue_2019'] or 0, reverse=reverse_sort)
            elif sort_criteria == 'profit':
                sorted_companies = sorted(companies, key=lambda x: x['profit_2019'] or 0, reverse=reverse_sort)
            elif sort_criteria == 'personnel':
                sorted_companies = sorted(companies, key=lambda x: x['personnel_2019'] or 0, reverse=reverse_sort)
            else:
                sorted_companies = companies
            
            # Convert to JSON format with all fields
            companies_data = []
            year_source = data.get('year_source', '2025')
            source_text = f"Україна {year_source}"
            
            for rank, company in enumerate(sorted_companies, 1):
                companies_data.append({
                    'ranking': rank,
                    'source': source_text,
                    'edrpou': company['edrpou'],
                    'company_name': company['name'],
                    'kved_code': company['kved_code'],
                    'kved_description': company['kved_description'],
                    'personnel_2019': company['personnel_2019'],
                    'region_name': company['region_name'],
                    'phone': company['phone'],
                    'address': company['address'],
                    'revenue_2019': company['revenue_2019'],
                    'profit_2019': company['profit_2019'],
                    'company_size_name': company['company_size_name'],
                    'first_name': company['first_name'],
                    'middle_name': company['middle_name'],
                    'last_name': company['last_name'],
                    'work_phone': company['work_phone'],
                    'corporate_site': company['corporate_site'],
                    'work_email': company['work_email'],
                    'company_status': company['company_status'],
                    'director': company['director'],
                    'government_purchases': company['government_purchases'],
                    'tender_count': company['tender_count'],
                    'initials': company['initials'],
                    'actualized': company['actualized']
                })
            
            # Helper function to get sort value (now for dictionary objects)
            def getCompanySortValue(company, criteria):
                if criteria == 'revenue':
                    return company['revenue_2019'] or 0
                elif criteria == 'profit':
                    return company['profit_2019'] or 0
                elif criteria == 'personnel':
                    return company['personnel_2019'] or 0
                return 0
            
            # Save ranking to database 
            # Get human-readable criteria name
            criteria_names = {
                'revenue': 'Чистий дохід від реалізації',
                'profit': 'Чистий фінансовий результат',
                'personnel': 'Кількість працівників'
            }
            criteria_display = criteria_names.get(sort_criteria, sort_criteria)
            
            # Get ranking name from form data
            ranking_name = data['ranking_name']
            
            # Generate source name (like the table source column format)
            from datetime import datetime
            current_year = datetime.now().year
            source_name = f"Україна {current_year}"
            
            logging.info(f"Creating ranking with data: {data}")
            
            # Use the same SelectionBase that we got earlier (no need to re-query)
            logging.info(f"Using SelectionBase ID: {selection_base.id} for ranking creation")
            
            # Create Ranking record with minimal required fields
            try:
                logging.info(f"Creating Ranking with name: {ranking_name}")
                ranking = Ranking(
                    name=ranking_name,
                    selection_base_id=selection_base.id,
                    companies_count=len(sorted_companies),
                    is_active=True
                )
                db.session.add(ranking)
                db.session.flush()  # Get ranking ID
                logging.info(f"Ranking created with ID: {ranking.id}")
            except Exception as e:
                logging.error(f"Error creating Ranking: {e}")
                raise
            
            # Clear all rankings first using raw SQL to avoid SQLAlchemy column issues
            try:
                logging.info("Clearing existing rankings")
                from sqlalchemy import text
                db.session.execute(text("UPDATE companies SET ranking = NULL, ranking_criteria = NULL"))
                logging.info("Existing rankings cleared")
            except Exception as e:
                logging.error(f"Error clearing rankings: {e}")
                raise
            
            # Create RankingCompany records and update companies table
            logging.info(f"Processing {len(sorted_companies)} companies for ranking")
            for rank, company in enumerate(sorted_companies, 1):
                try:
                    # Create RankingCompany record
                    ranking_company = RankingCompany(
                        ranking_id=ranking.id,
                        company_id=company['id'],
                        position=rank
                    )
                    db.session.add(ranking_company)
                    
                    # Update current ranking in companies table
                    db.session.execute(text(
                        "UPDATE companies SET ranking = :rank, ranking_criteria = :criteria WHERE id = :company_id"
                    ), {
                        'rank': rank,
                        'criteria': criteria_display,
                        'company_id': company['id']
                    })
                    
                    # Save to ranking history
                    db.session.execute(text(
                        "INSERT INTO company_ranking_history (company_id, ranking_name, ranking_position, ranking_criteria, source_name) VALUES (:company_id, :name, :position, :criteria, :source_name)"
                    ), {
                        'company_id': company['id'],
                        'name': ranking_name,
                        'position': rank,
                        'criteria': criteria_display,
                        'source_name': source_name
                    })
                    
                    if rank % 100 == 0:  # Log every 100 companies
                        logging.info(f"Processed {rank} companies")
                        
                except Exception as e:
                    logging.error(f"Error processing company {company['id']} at rank {rank}: {e}")
                    raise
            
            try:
                logging.info("Committing transaction")
                db.session.commit()
                logging.info("Transaction committed successfully")
            except Exception as e:
                logging.error(f"Error committing transaction: {e}")
                raise
            
            return jsonify({
                'success': True, 
                'companies': companies_data,
                'ranking_id': ranking.id,  # Add ranking ID for PDF export
                'count': len(companies_data),
                'message': f'Створено рейтинг "{data["ranking_name"]}" з {len(companies_data)} компаній'
            })
            
        except Exception as e:
            import traceback
            logging.error(f"Error creating ranking: {e}")
            logging.error(f"Full traceback: {traceback.format_exc()}")
            db.session.rollback()  # Rollback on error
            return jsonify({'success': False, 'error': f'Помилка при створенні рейтингу: {str(e)}'})
    
    # GET request - show new ranking page
    return render_template('ranking.html')

@main.route('/export', methods=['GET', 'POST'])
@login_required
@manager_or_admin_required
def export():
    
    if request.method == 'POST':
        # Stage 3: Create final ranking with additional parameters
        sort_criteria = request.form.get('sort_criteria', 'revenue')
        year_source = request.form.get('year_source', '2025')
        ranking_name = request.form.get('ranking_name', 'Новий рейтинг')
        
        # Additional filters for ranking stage
        selected_regions = request.form.getlist('region_filter')
        selected_kved = request.form.getlist('kved_filter')
        selected_sizes = request.form.getlist('size_filter')
        
        try:
            # Get companies from selection base (stored in session)
            selected_ids = session.get('selected_company_ids', [])
            if not selected_ids:
                flash('Спочатку створіть базу відбору в розділі "Відбір компаній для рейтингу"', 'warning')
                return redirect(url_for('main.filter_companies_route'))
            
            # Start with selected companies
            query = db.select(Company).where(Company.id.in_(selected_ids))
            
            # Apply additional ranking-stage filters if any
            filters = []
            if selected_regions:
                filters.append(Company.region_name.in_(selected_regions))
            if selected_kved:
                filters.append(Company.kved_code.in_(selected_kved))
            if selected_sizes:
                filters.append(Company.company_size_name.in_(selected_sizes))
                
            if filters:
                from sqlalchemy import and_
                query = query.where(and_(*filters))
            
            # Execute query
            ranking_companies = db.session.execute(query).scalars().all()
            
            # Sort by selected criteria
            if sort_criteria == 'revenue':
                sorted_companies = sorted(ranking_companies, key=lambda x: x.revenue_2019 or 0, reverse=True)
            elif sort_criteria == 'profit':
                sorted_companies = sorted(ranking_companies, key=lambda x: x.profit_2019 or 0, reverse=True)
            elif sort_criteria == 'personnel':
                sorted_companies = sorted(ranking_companies, key=lambda x: x.personnel_2019 or 0, reverse=True)
            else:
                sorted_companies = ranking_companies
                
            # Clear all rankings first
            db.session.execute(db.update(Company).values(ranking=None))
            
            # Assign new rankings
            for rank, company in enumerate(sorted_companies, 1):
                company.ranking = rank
                
            db.session.commit()
            
            flash(f'Створено рейтинг "{ranking_name}" з {len(sorted_companies)} компаній за критерієм "{sort_criteria}"', 'success')
            
            # Update session with new ranking info
            session['last_ranking_name'] = ranking_name
            session['last_ranking_count'] = len(sorted_companies)
            session['last_ranking_criteria'] = sort_criteria
            
            # Show results immediately with table displayed
            flash('Рейтинг створено успішно!', 'success')
            return render_template('ranking.html',
                                 companies_with_rankings=sorted_companies,
                                 total_companies=db.session.execute(db.select(db.func.count(Company.id))).scalar() or 0,
                                 selection_count=session.get('selection_count', 0),
                                 selection_criteria=session.get('selection_criteria', 'Не визначено'),
                                 ranked_companies=len(sorted_companies),
                                 ranking_name=ranking_name,
                                 sort_criteria=sort_criteria,
                                 year_source=year_source,
                                 top_by_revenue=sorted_companies[:5],
                                 top_by_profit=sorted_companies[:5], 
                                 top_by_personnel=sorted_companies[:5],
                                 show_table=True)
                                 
        except Exception as e:
            logging.error(f"Error creating ranking: {e}")
            flash('Помилка при створенні рейтингу.', 'danger')
    
    # GET request - show ranking creation form
    try:
        # Get selection base info
        selected_ids = session.get('selected_company_ids', [])
        selection_criteria = session.get('selection_criteria', 'Не визначено')
        selection_count = session.get('selection_count', 0)
        selection_info = session.get('selection_info', None)
        
        # Get companies that already have rankings
        companies_with_rankings = db.session.execute(
            db.select(Company)
            .where(Company.ranking.isnot(None))
            .order_by(asc(Company.ranking))
        ).scalars().all()
        
        # Get statistics
        total_companies = db.session.execute(db.select(db.func.count(Company.id))).scalar() or 0
        
        return render_template('ranking.html',
                             companies_with_rankings=companies_with_rankings,
                             total_companies=total_companies,
                             selection_count=selection_count,
                             selection_criteria=selection_criteria,
                             selection_info=selection_info,
                             ranked_companies=len(companies_with_rankings),
                             top_by_revenue=companies_with_rankings[:5],
                             top_by_profit=companies_with_rankings[:5],
                             top_by_personnel=companies_with_rankings[:5])
        
    except Exception as e:
        logging.error(f"Error in ranking page: {str(e)}")
        flash('Помилка завантаження сторінки рейтингу.', 'danger')
        return redirect(url_for('main.index'))

@main.route('/ranking/<int:ranking_id>/export-pdf')
@login_required  
def export_ranking_pdf(ranking_id):
    """Експорт рейтингу в PDF з фірмовим бланком"""
    if not current_user.has_permission('edit'):
        flash('У вас немає прав для експорту рейтингів.', 'danger')
        return redirect(url_for('main.companies'))
    
    try:
        from pdf_export import export_ranking_to_pdf
        success, filename, message = export_ranking_to_pdf(ranking_id)
        
        if success:
            # Повертаємо файл для завантаження
            file_path = os.path.join('static/exports', filename)
            return send_file(
                file_path,
                as_attachment=True,
                download_name=f'ranking_{ranking_id}.pdf',
                mimetype='application/pdf'
            )
        else:
            flash(f'Помилка експорту: {message}', 'danger')
            return redirect(url_for('main.companies'))
            
    except Exception as e:
        flash(f'Помилка створення PDF: {str(e)}', 'danger')
        return redirect(url_for('main.companies'))

@main.route('/api/latest-ranking-id')
@login_required
def get_latest_ranking_id():
    """Get latest ranking ID for PDF export"""
    try:
        latest_ranking = db.session.execute(
            db.select(Ranking).order_by(Ranking.created_at.desc())
        ).scalar_one_or_none()
        
        if latest_ranking:
            return jsonify({'success': True, 'ranking_id': latest_ranking.id})
        else:
            return jsonify({'success': False, 'error': 'Немає доступних рейтингів'})
            
    except Exception as e:
        logging.error(f"Error getting latest ranking: {e}")
        return jsonify({'success': False, 'error': str(e)})

@main.route('/api/current-selection-info')
@login_required
def get_current_selection_info():
    """API endpoint для отримання актуальної інформації про поточний відбір"""
    try:
        # Загальна кількість компаній в базі
        total_companies = db.session.execute(db.select(db.func.count(Company.id))).scalar() or 0
        
        # Отримуємо поточний відбір з бази даних (не з session)
        active_selection = db.session.execute(
            db.select(SelectionBase).where(SelectionBase.is_active == True)
        ).scalar_one_or_none()
        
        selection_count = active_selection.companies_count if active_selection else 0
        
        # Кількість проранжованих компаній
        ranked_companies_count = db.session.execute(
            db.select(db.func.count(Company.id))
            .where(Company.ranking.isnot(None))
        ).scalar() or 0
        
        # Критерії відбору з бази даних
        if active_selection:
            criteria_parts = []
            if active_selection.min_employees:
                criteria_parts.append(f"Працівників ≥ {active_selection.min_employees}")
            if active_selection.min_revenue:
                criteria_parts.append(f"Дохід ≥ {active_selection.min_revenue:,.0f}")
            if active_selection.min_profit:
                criteria_parts.append(f"Прибуток ≥ {active_selection.min_profit:,.0f}")
            selection_criteria = " | ".join(criteria_parts) if criteria_parts else "Критерії не задано"
        else:
            selection_criteria = 'Критерії не задано'
        
        # Детальна інформація про поточний відбір
        selection_info = None
        if active_selection:
            selection_info = {
                'min_personnel': active_selection.min_employees,
                'min_revenue': float(active_selection.min_revenue) if active_selection.min_revenue else 0,
                'min_profit': float(active_selection.min_profit) if active_selection.min_profit else None,
                'companies_count': active_selection.companies_count
            }
        
        return jsonify({
            'success': True,
            'total_in_database': total_companies,
            'total_in_selection': selection_count,
            'companies_with_ranking': ranked_companies_count,
            'selection_criteria': selection_criteria,
            'selection_info': selection_info
        })
        
    except Exception as e:
        logging.error(f"Error getting current selection info: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ===== УПРАВЛІННЯ КОРИСТУВАЧАМИ (тільки адміністратори) =====

@main.route('/users')
@login_required
@require_role('admin')
def users():
    """Сторінка управління користувачами"""
    try:
        users = db.session.execute(db.select(User).order_by(User.created_at.desc())).scalars().all()
        return render_template('users.html', users=users)
    except Exception as e:
        logging.error(f"Error loading users: {e}")
        flash('Помилка завантаження користувачів', 'danger')
        return redirect(url_for('main.index'))

@main.route('/users/create', methods=['POST'])
@login_required
@require_role('admin')
def create_user():
    """Створення нового користувача"""
    try:
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        role = request.form.get('role', 'guest')
        
        # Валідація
        errors = []
        
        if not username or len(username) < 3:
            errors.append('Ім\'я користувача має бути не менше 3 символів.')
        elif db.session.execute(db.select(User).where(User.username == username)).scalar_one_or_none():
            errors.append('Користувач з таким іменем вже існує.')
            
        if not email or '@' not in email:
            errors.append('Введіть коректний email.')
        elif db.session.execute(db.select(User).where(User.email == email)).scalar_one_or_none():
            errors.append('Користувач з таким email вже існує.')
            
        if not password or len(password) < 6:
            errors.append('Пароль має бути не менше 6 символів.')
            
        if role not in ['admin', 'manager', 'guest']:
            errors.append('Некоректна роль.')
        
        if errors:
            for error in errors:
                flash(error, 'danger')
            return redirect(url_for('main.users'))
        
        # Створення користувача
        user = User()
        user.username = username
        user.email = email
        user.password_hash = generate_password_hash(password)
        user.role = role
        
        db.session.add(user)
        db.session.commit()
        
        flash(f'Користувача "{username}" успішно створено!', 'success')
        return redirect(url_for('main.users'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error creating user: {e}")
        flash('Помилка створення користувача', 'danger')
        return redirect(url_for('main.users'))

@main.route('/users/<int:user_id>/edit', methods=['POST'])
@login_required
@require_role('admin')
def edit_user(user_id):
    """Редагування користувача"""
    try:
        user = db.session.execute(db.select(User).where(User.id == user_id)).scalar_one_or_none()
        if not user:
            flash('Користувача не знайдено', 'danger')
            return redirect(url_for('main.users'))
        
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'guest')
        
        # Валідація
        errors = []
        
        if not username or len(username) < 3:
            errors.append('Ім\'я користувача має бути не менше 3 символів.')
        elif username != user.username and db.session.execute(db.select(User).where(User.username == username)).scalar_one_or_none():
            errors.append('Користувач з таким іменем вже існує.')
            
        if not email or '@' not in email:
            errors.append('Введіть коректний email.')
        elif email != user.email and db.session.execute(db.select(User).where(User.email == email)).scalar_one_or_none():
            errors.append('Користувач з таким email вже існує.')
            
        if password and len(password) < 6:
            errors.append('Пароль має бути не менше 6 символів.')
            
        if role not in ['admin', 'manager', 'guest']:
            errors.append('Некоректна роль.')
        
        if errors:
            for error in errors:
                flash(error, 'danger')
            return redirect(url_for('main.users'))
        
        # Оновлення користувача
        user.username = username
        user.email = email
        if password:  # Змінюємо пароль тільки якщо введено новий
            user.password_hash = generate_password_hash(password)
        user.role = role
        
        db.session.commit()
        
        flash(f'Користувача "{username}" успішно оновлено!', 'success')
        return redirect(url_for('main.users'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error editing user: {e}")
        flash('Помилка редагування користувача', 'danger')
        return redirect(url_for('main.users'))

@main.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@require_role('admin')
def delete_user(user_id):
    """Видалення користувача"""
    try:
        user = db.session.execute(db.select(User).where(User.id == user_id)).scalar_one_or_none()
        if not user:
            flash('Користувача не знайдено', 'danger')
            return redirect(url_for('main.users'))
        
        # Не дозволяємо видаляти самого себе
        if user.id == current_user.id:
            flash('Не можна видалити власний аккаунт', 'warning')
            return redirect(url_for('main.users'))
        
        # Не дозволяємо видаляти головного адміна
        if user.username == 'admin':
            flash('Не можна видалити головного адміністратора', 'warning')
            return redirect(url_for('main.users'))
        
        username = user.username
        db.session.delete(user)
        db.session.commit()
        
        flash(f'Користувача "{username}" успішно видалено!', 'success')
        return redirect(url_for('main.users'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error deleting user: {e}")
        flash('Помилка видалення користувача', 'danger')
        return redirect(url_for('main.users'))

@main.route('/debug/actualization')
@login_required
@require_role('admin')
def debug_actualization():
    """Діагностика даних актуалізації в production базі"""
    try:
        # Показати структуру таблиці companies
        structure_query = text("""
            SELECT column_name, data_type, is_nullable 
            FROM information_schema.columns 
            WHERE table_name = 'companies' 
            AND column_name LIKE '%actual%'
            ORDER BY ordinal_position;
        """)
        
        # Приклади значень актуалізації
        samples_query = text("""
            SELECT DISTINCT actualized, COUNT(*) as count
            FROM companies 
            WHERE actualized IS NOT NULL 
            AND actualized != ''
            GROUP BY actualized
            ORDER BY count DESC
            LIMIT 20;
        """)
        
        # Загальна статистика
        stats_query = text("""
            SELECT 
                COUNT(*) as total_companies,
                COUNT(actualized) as with_actualized_field,
                COUNT(CASE WHEN actualized IS NOT NULL AND actualized != '' THEN 1 END) as with_actualized_value
            FROM companies;
        """)
        
        structure = db.session.execute(structure_query).fetchall()
        samples = db.session.execute(samples_query).fetchall()
        stats = db.session.execute(stats_query).fetchone()
        
        debug_info = {
            'structure': [{'column': row[0], 'type': row[1], 'nullable': row[2]} for row in structure],
            'samples': [{'value': row[0], 'count': row[1]} for row in samples],
            'stats': {
                'total': stats[0],
                'with_field': stats[1], 
                'with_value': stats[2]
            } if stats else {}
        }
        
        return f"""
        <h2>🔍 Діагностика актуалізації</h2>
        <h3>📊 Статистика:</h3>
        <p>Всього компаній: {debug_info['stats'].get('total', 0)}</p>
        <p>З полем actualized: {debug_info['stats'].get('with_field', 0)}</p>
        <p>З значенням в actualized: {debug_info['stats'].get('with_value', 0)}</p>
        
        <h3>🗂️ Структура поля actualized:</h3>
        <ul>{''.join([f"<li>{s['column']} - {s['type']} - nullable: {s['nullable']}</li>" for s in debug_info['structure']])}</ul>
        
        <h3>📋 Приклади значень:</h3>
        <ul>{''.join([f"<li>'{s['value']}' - {s['count']} разів</li>" for s in debug_info['samples']])}</ul>
        
        <a href="/">← Назад на головну</a>
        """
    
    except Exception as e:
        return f"Помилка діагностики: {str(e)}"

@main.route('/debug/structure')
@login_required 
@require_role('admin')
def debug_structure():
    """Показати реальну структуру таблиці companies"""
    try:
        structure_query = text("""
            SELECT column_name, data_type, is_nullable, column_default
            FROM information_schema.columns 
            WHERE table_name = 'companies' 
            ORDER BY ordinal_position;
        """)
        
        columns = db.session.execute(structure_query).fetchall()
        
        result = "<h2>🏗️ Реальна структура таблиці companies в production</h2><table border='1' style='border-collapse: collapse;'>"
        result += "<tr><th>Колонка</th><th>Тип</th><th>Nullable</th><th>Default</th></tr>"
        
        for col in columns:
            result += f"<tr><td>{col[0]}</td><td>{col[1]}</td><td>{col[2]}</td><td>{col[3] or ''}</td></tr>"
        
        result += "</table><a href='/'>← Головна</a>"
        return result
    
    except Exception as e:
        return f"Помилка: {str(e)}"

@main.route('/debug/company/<edrpou>')
@login_required 
@require_role('admin')
def debug_company(edrpou):
    """Діагностика конкретної компанії за ЄДРПОУ"""
    try:
        # Спочатку знайти компанію з мінімальним набором полів
        basic_query = text("""
            SELECT id, edrpou, name, actualized
            FROM companies 
            WHERE edrpou = :edrpou
        """)
        
        result = db.session.execute(basic_query, {'edrpou': edrpou}).fetchone()
        
        if not result:
            return f"<h2>❌ Компанію з ЄДРПОУ {edrpou} не знайдено</h2><a href='/'>← Назад</a>"
        
        # Детальний аналіз поля actualized
        detailed_query = text("""
            SELECT 
                actualized,
                char_length(actualized) as actualized_length,
                CASE WHEN ascii(actualized) IS NOT NULL THEN ascii(actualized) ELSE 0 END as actualized_ascii,
                CASE 
                    WHEN actualized IS NULL THEN 'NULL'
                    WHEN actualized = '' THEN 'EMPTY'
                    WHEN UPPER(TRIM(actualized)) = 'НІ' THEN 'НІ (uppercase)'
                    WHEN LOWER(TRIM(actualized)) = 'ні' THEN 'ні (lowercase)'
                    WHEN TRIM(actualized) = 'так' THEN 'так'
                    WHEN TRIM(actualized) = 'ТАК' THEN 'ТАК'
                    ELSE CONCAT('OTHER: "', actualized, '"')
                END as actualized_analysis
            FROM companies 
            WHERE edrpou = :edrpou
        """)
        
        detail = db.session.execute(detailed_query, {'edrpou': edrpou}).fetchone()
        
        return f"""
        <h2>🔍 Діагностика компанії {edrpou}</h2>
        <table border="1" style="border-collapse: collapse;">
        <tr><td><b>ID</b></td><td>{result[0]}</td></tr>
        <tr><td><b>ЄДРПОУ</b></td><td>{result[1]}</td></tr>
        <tr><td><b>Назва</b></td><td>{result[2]}</td></tr>
        <tr><td><b>Поле actualized</b></td><td>'{detail[0] if detail else result[3]}'</td></tr>
        <tr><td><b>Довжина actualized</b></td><td>{detail[1] if detail else 'N/A'}</td></tr>
        <tr><td><b>ASCII код першого символу</b></td><td>{detail[2] if detail else 'N/A'}</td></tr>
        <tr><td><b>Аналіз actualized</b></td><td>{detail[3] if detail else 'N/A'}</td></tr>
        </table>
        
        <h3>🧪 Тест поточної логіки:</h3>
        <p>Поточна логіка вважає актуалізованими: НЕ NULL, НЕ пусте, НЕ 'НІ', 'NO', 'FALSE', '0', 'NULL', 'NONE'</p>
        <p><b>Результат для цієї компанії:</b> 
        {"✅ АКТУАЛІЗОВАНО" if result[3] and result[3].strip() and result[3].strip().upper() not in ['НІ', 'NO', 'FALSE', '0', 'NULL', 'NONE'] else "❌ НЕ АКТУАЛІЗОВАНО"}
        </p>
        
        <a href="/debug/structure">← Структура таблиці</a> | 
        <a href="/">← Головна</a>
        """
    
    except Exception as e:
        return f"Помилка діагностики компанії: {str(e)}"
